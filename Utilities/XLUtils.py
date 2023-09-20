import openpyxl
from openpyxl.styles import PatternFill


class XL_operations:

    def GetRows_count(self, path, sheetname):
        workbook = openpyxl.load_workbook(path)
        sheetnames = workbook.sheetnames
        # print(sheetnames)
        # # we can also get the name of the sheet which is active by using the active property
        # print("Active sheet: ", workbook.active)
        sheet = workbook[sheetname]
        return sheet.max_row

    def GetColumns_count(self, path, sheetname):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        return sheet.max_column

    def ReadData_from_XL(self, path, sheetname, row, column):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        return sheet.cell(row, column).value

    def WriteData_from_XL(self, path, sheetname, row, column, data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        sheet.cell(row, column).value = None
        sheet.cell(row, column).value = data
        workbook.save(path)

    def Fill_redencolor(self, path, sheetname, row, column):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        redFill = PatternFill(start_color='FF0000',
                              end_color='FF0000',
                              fill_type="solid")  # used hex code for red color

        sheet.cell(row, column).fill = redFill
        workbook.save(path)

    def Fill_greenencolor(self, path, sheetname, row, column):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]
        greenFill = PatternFill(start_color='60b212',
                                end_color='60b212',
                                fill_type="solid")  # used hex code for red color

        sheet.cell(row, column).fill = greenFill
        workbook.save(path)
